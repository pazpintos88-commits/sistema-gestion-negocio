from werkzeug.security import generate_password_hash, check_password_hash
import io
import os
import sqlite3
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)
app.secret_key = 'clave_secreta_super_segura_para_el_sistema'
DATABASE = 'negocio.db'


# ==========================================
# BASE DE DATOS Y CONEXIÓN
# ==========================================


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DATABASE)
    cursor = db.cursor()

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS configuracion (
        id INTEGER PRIMARY KEY DEFAULT 1,
        nombre_negocio TEXT DEFAULT 'Mi Negocio',
        cuit TEXT DEFAULT '',
        direccion TEXT DEFAULT ''
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        usuario TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        rol TEXT NOT NULL DEFAULT 'vendedor',
        activo INTEGER DEFAULT 1
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        categoria TEXT DEFAULT 'General',
        costo REAL NOT NULL DEFAULT 0,
        precio_venta REAL NOT NULL DEFAULT 0,
        stock INTEGER NOT NULL DEFAULT 0,
        stock_min INTEGER NOT NULL DEFAULT 3
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS ventas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        vendedor_id INTEGER,
        total REAL NOT NULL,
        FOREIGN KEY (vendedor_id) REFERENCES usuarios (id)
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS venta_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        venta_id INTEGER NOT NULL,
        producto_id INTEGER NOT NULL,
        nombre_producto TEXT NOT NULL,
        categoria_producto TEXT DEFAULT 'General',
        costo_unitario REAL NOT NULL,
        precio_unitario REAL NOT NULL,
        cantidad INTEGER NOT NULL,
        subtotal REAL NOT NULL,
        FOREIGN KEY (venta_id) REFERENCES ventas (id),
        FOREIGN KEY (producto_id) REFERENCES productos (id)
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS gastos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        descripcion TEXT NOT NULL,
        categoria TEXT DEFAULT 'General',
        monto REAL NOT NULL
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS facturas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        venta_id INTEGER UNIQUE NOT NULL,
        numero TEXT UNIQUE NOT NULL,
        fecha TEXT NOT NULL,
        cliente_nombre TEXT DEFAULT 'Consumidor Final',
        cliente_doc TEXT DEFAULT '',
        notas TEXT DEFAULT '',
        total REAL NOT NULL,
        FOREIGN KEY (venta_id) REFERENCES ventas (id)
    )
    ''')

    cursor.execute('SELECT COUNT(*) FROM configuracion')
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            'INSERT INTO configuracion (id, nombre_negocio) VALUES (1, "Mi Negocio")'
        )

    cursor.execute('SELECT COUNT(*) FROM usuarios')
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            'INSERT INTO usuarios (nombre, usuario, password_hash, rol) VALUES (?, ?, ?, ?)',
            (
                'Administrador',
                'admin',
                generate_password_hash('admin123'),
                'admin',
            ),
        )

    db.commit()
    db.close()


# ==========================================
# FILTROS Y MIDDLEWARES
# ==========================================


@app.template_filter('money')
def money_filter(value):
    try:
        return f'${float(value):,.2f}'
    except (ValueError, TypeError):
        return '$0.00'


@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        db = get_db()
        g.user = db.execute(
            'SELECT * FROM usuarios WHERE id = ? AND activo = 1', (user_id,)
        ).fetchone()

    if g.user:
        db = get_db()
        g.config = db.execute(
            'SELECT * FROM configuracion WHERE id = 1'
        ).fetchone()


def login_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if g.user is None:
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if g.user is None or g.user['rol'] != 'admin':
            return render_template(
                'error.html',
                mensaje='Solo los administradores pueden ingresar a esta sección.',
            )
        return f(*args, **kwargs)

    return decorated_function


# ==========================================
# AUTENTICACIÓN
# ==========================================


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']
        db = get_db()
        user = db.execute(
            'SELECT * FROM usuarios WHERE usuario = ? AND activo = 1', (usuario,)
        ).fetchone()

        if user and check_password_hash(user['password'], password):
            session.clear()
            session['user_id'] = user['id']
            session['usuario'] = user['usuario']
            session['rol'] = user['rol']
            return redirect(url_for('index'))
        
        flash('Usuario o contraseña incorrectos', 'danger')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']
        
        db = get_db()
        error = None

        if not usuario or not password:
            error = 'Usuario y contraseña son requeridos.'
        elif db.execute('SELECT id FROM usuarios WHERE usuario = ?', (usuario,)).fetchone() is not None:
            error = f'El usuario {usuario} ya está registrado.'

        if error is None:
            hashed_pw = generate_password_hash(password)
            db.execute(
                'INSERT INTO usuarios (usuario, password, rol, activo) VALUES (?, ?, ?, ?)',
                (usuario, hashed_pw, 'admin', 1)
            )
            db.commit()
            flash('¡Cuenta creada con éxito! Ahora puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))

        flash(error, 'danger')

    return render_template('register.html')        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id'] = user['id']
            return redirect(url_for('dashboard'))
        else:
            return render_template(
                'login.html', error='Usuario o contraseña incorrectos.'
            )

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ==========================================
# DASHBOARD
# ==========================================


@app.route('/')
@login_required
def dashboard():
    db = get_db()
    rango = request.args.get('rango', 'hoy')

    hoy = datetime.now()
    filtro_fecha_ventas = ''
    filtro_fecha_gastos = ''
    params_ventas = []
    params_gastos = []

    if rango == 'hoy':
        inicio = hoy.strftime('%Y-%m-%d 00:00:00')
        fin = hoy.strftime('%Y-%m-%d 23:59:59')
        filtro_fecha_ventas = 'WHERE fecha BETWEEN ? AND ?'
        filtro_fecha_gastos = 'WHERE fecha BETWEEN ? AND ?'
        params_ventas = [inicio, fin]
        params_gastos = [hoy.strftime('%Y-%m-%d'), hoy.strftime('%Y-%m-%d')]
        etiqueta_rango = 'Hoy'
    elif rango == '7':
        inicio = (hoy - timedelta(days=7)).strftime('%Y-%m-%d 00:00:00')
        filtro_fecha_ventas = 'WHERE fecha >= ?'
        filtro_fecha_gastos = 'WHERE fecha >= ?'
        params_ventas = [inicio]
        params_gastos = [(hoy - timedelta(days=7)).strftime('%Y-%m-%d')]
        etiqueta_rango = 'Últimos 7 Días'
    elif rango == '30':
        inicio = (hoy - timedelta(days=30)).strftime('%Y-%m-%d 00:00:00')
        filtro_fecha_ventas = 'WHERE fecha >= ?'
        filtro_fecha_gastos = 'WHERE fecha >= ?'
        params_ventas = [inicio]
        params_gastos = [(hoy - timedelta(days=30)).strftime('%Y-%m-%d')]
        etiqueta_rango = 'Últimos 30 Días'
    else:
        etiqueta_rango = 'Histórico Total'

    tot_ventas = db.execute(
        f'SELECT SUM(total) FROM ventas {filtro_fecha_ventas}', params_ventas
    ).fetchone()[0] or 0

    if filtro_fecha_ventas:
        tot_costo = db.execute(
            f'SELECT SUM(vi.costo_unitario * vi.cantidad) FROM venta_items vi JOIN ventas v ON vi.venta_id = v.id {filtro_fecha_ventas}',
            params_ventas,
        ).fetchone()[0] or 0
    else:
        tot_costo = db.execute(
            'SELECT SUM(costo_unitario * cantidad) FROM venta_items'
        ).fetchone()[0] or 0

    tot_gastos = db.execute(
        f'SELECT SUM(monto) FROM gastos {filtro_fecha_gastos}', params_gastos
    ).fetchone()[0] or 0

    ganancia_neta = tot_ventas - tot_costo - tot_gastos

    bajos = db.execute(
        'SELECT * FROM productos WHERE stock <= stock_min'
    ).fetchall()

    return render_template(
        'dashboard.html',
        total_ventas=tot_ventas,
        total_costo=tot_costo,
        total_gastos=tot_gastos,
        ganancia_neta=ganancia_neta,
        bajos=bajos,
        rango=rango,
        etiqueta_rango=etiqueta_rango,
    )


# ==========================================
# PRODUCTOS
# ==========================================


@app.route('/productos', methods=['GET', 'POST'])
@login_required
def productos():
    db = get_db()
    if request.method == 'POST':
        if g.user['rol'] != 'admin':
            flash('Solo los administradores pueden agregar productos.', 'error')
            return redirect(url_for('productos'))

        nombre = request.form['nombre']
        categoria = request.form.get('categoria', 'General')
        costo = float(request.form.get('costo', 0))
        precio_venta = float(request.form.get('precio_venta', 0))
        stock = int(request.form.get('stock', 0))
        stock_min = int(request.form.get('stock_min', 3))

        db.execute(
            'INSERT INTO productos (nombre, categoria, costo, precio_venta, stock, stock_min) VALUES (?, ?, ?, ?, ?, ?)',
            (nombre, categoria, costo, precio_venta, stock, stock_min),
        )
        db.commit()
        flash('Producto agregado con éxito.', 'success')
        return redirect(url_for('productos'))

    prods = db.execute('SELECT * FROM productos ORDER BY nombre').fetchall()
    return render_template('productos.html', productos=prods)


@app.route('/productos/eliminar/<int:pid>', methods=['POST'])
@admin_required
def eliminar_producto(pid):
    db = get_db()
    db.execute('DELETE FROM productos WHERE id = ?', (pid,))
    db.commit()
    flash('Producto eliminado.', 'success')
    return redirect(url_for('productos'))


# ==========================================
# VENTAS
# ==========================================


@app.route('/ventas')
@login_required
def ventas():
    db = get_db()
    v = db.execute('''
        SELECT v.*, u.nombre as vendedor 
        FROM ventas v 
        LEFT JOIN usuarios u ON v.vendedor_id = u.id 
        ORDER BY v.id DESC
    ''').fetchall()

    facturadas = [
        f['venta_id']
        for f in db.execute('SELECT venta_id FROM facturas').fetchall()
    ]

    return render_template('ventas.html', ventas=v, facturadas=facturadas)


@app.route('/ventas/nueva')
@login_required
def venta_nueva():
    db = get_db()
    prods = db.execute(
        'SELECT * FROM productos WHERE stock > 0 ORDER BY nombre'
    ).fetchall()
    carrito = session.get('carrito', [])
    total = sum(item['subtotal'] for item in carrito)
    return render_template(
        'venta_nueva.html', productos=prods, carrito=carrito, total=total
    )


@app.route('/ventas/agregar-item', methods=['POST'])
@login_required
def venta_agregar_item():
    producto_id = int(request.form['producto_id'])
    cantidad = int(request.form['cantidad'])

    db = get_db()
    prod = db.execute(
        'SELECT * FROM productos WHERE id = ?', (producto_id,)
    ).fetchone()

    if not prod or prod['stock'] < cantidad:
        flash('Stock insuficiente para este producto.', 'error')
        return redirect(url_for('venta_nueva'))

    carrito = session.get('carrito', [])

    for item in carrito:
        if item['producto_id'] == producto_id:
            if item['cantidad'] + cantidad > prod['stock']:
                flash('Superas el stock disponible.', 'error')
                return redirect(url_for('venta_nueva'))
            item['cantidad'] += cantidad
            item['subtotal'] = item['cantidad'] * item['precio_unitario']
            session['carrito'] = carrito
            return redirect(url_for('venta_nueva'))

    carrito.append({
        'producto_id': prod['id'],
        'nombre': prod['nombre'],
        'categoria': prod['categoria'],
        'costo_unitario': prod['costo'],
        'precio_unitario': prod['precio_venta'],
        'cantidad': cantidad,
        'subtotal': cantidad * prod['precio_venta'],
    })

    session['carrito'] = carrito
    return redirect(url_for('venta_nueva'))


@app.route('/ventas/quitar-item/<int:idx>', methods=['POST'])
@login_required
def venta_quitar_item(idx):
    carrito = session.get('carrito', [])
    if 0 <= idx < len(carrito):
        carrito.pop(idx)
        session['carrito'] = carrito
    return redirect(url_for('venta_nueva'))


@app.route('/ventas/confirmar', methods=['POST'])
@login_required
def venta_confirmar():
    carrito = session.get('carrito', [])
    if not carrito:
        flash('El carrito está vacío.', 'error')
        return redirect(url_for('venta_nueva'))

    db = get_db()
    fecha_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    total_venta = sum(item['subtotal'] for item in carrito)

    cur = db.cursor()
    cur.execute(
        'INSERT INTO ventas (fecha, vendedor_id, total) VALUES (?, ?, ?)',
        (fecha_actual, g.user['id'], total_venta),
    )
    venta_id = cur.lastrowid

    for item in carrito:
        cur.execute(
            '''
            INSERT INTO venta_items (venta_id, producto_id, nombre_producto, categoria_producto, costo_unitario, precio_unitario, cantidad, subtotal)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''',
            (
                venta_id,
                item['producto_id'],
                item['nombre'],
                item['categoria'],
                item['costo_unitario'],
                item['precio_unitario'],
                item['cantidad'],
                item['subtotal'],
            ),
        )

        cur.execute(
            'UPDATE productos SET stock = stock - ? WHERE id = ?',
            (item['cantidad'], item['producto_id']),
        )

    db.commit()
    session['carrito'] = []
    flash('Venta realizada con éxito.', 'success')
    return redirect(url_for('venta_detalle', vid=venta_id))


@app.route('/ventas/detalle/<int:vid>')
@login_required
def venta_detalle(vid):
    db = get_db()
    v = db.execute('''
        SELECT v.*, u.nombre as vendedor 
        FROM ventas v 
        LEFT JOIN usuarios u ON v.vendedor_id = u.id 
        WHERE v.id = ?
    ''', (vid,)).fetchone()

    if not v:
        flash('Venta no encontrada.', 'error')
        return redirect(url_for('ventas'))

    items = db.execute(
        'SELECT * FROM venta_items WHERE venta_id = ?', (vid,)
    ).fetchall()
    factura = db.execute(
        'SELECT * FROM facturas WHERE venta_id = ?', (vid,)
    ).fetchone()

    return render_template(
        'venta_detalle.html', venta=v, items=items, factura=factura
    )


# ==========================================
# GASTOS
# ==========================================


@app.route('/gastos', methods=['GET', 'POST'])
@admin_required
def gastos():
    db = get_db()
    if request.method == 'POST':
        descripcion = request.form['descripcion']
        categoria = request.form.get('categoria', 'General')
        monto = float(request.form['monto'])
        fecha = request.form.get(
            'fecha', datetime.now().strftime('%Y-%m-%d')
        )

        db.execute(
            'INSERT INTO gastos (fecha, descripcion, categoria, monto) VALUES (?, ?, ?, ?)',
            (fecha, descripcion, categoria, monto),
        )
        db.commit()
        flash('Gasto registrado con éxito.', 'success')
        return redirect(url_for('gastos'))

    lista_gastos = db.execute(
        'SELECT * FROM gastos ORDER BY id DESC'
    ).fetchall()
    hoy_str = datetime.now().strftime('%Y-%m-%d')
    return render_template(
        'gastos.html', gastos=lista_gastos, hoy=hoy_str
    )


@app.route('/gastos/eliminar/<int:gid>', methods=['POST'])
@admin_required
def eliminar_gasto(gid):
    db = get_db()
    db.execute('DELETE FROM gastos WHERE id = ?', (gid,))
    db.commit()
    flash('Gasto eliminado.', 'success')
    return redirect(url_for('gastos'))


# ==========================================
# FACTURAS Y PDF
# ==========================================


@app.route('/facturas')
@login_required
def facturas():
    db = get_db()
    f = db.execute('SELECT * FROM facturas ORDER BY id DESC').fetchall()
    return render_template('facturas.html', facturas=f)


@app.route('/facturas/nueva/<int:vid>', methods=['GET', 'POST'])
@login_required
def factura_nueva(vid):
    db = get_db()
    v = db.execute('SELECT * FROM ventas WHERE id = ?', (vid,)).fetchone()

    if not v:
        flash('Venta no encontrada.', 'error')
        return redirect(url_for('ventas'))

    if request.method == 'POST':
        cliente_nombre = request.form.get('cliente_nombre', 'Consumidor Final')
        cliente_doc = request.form.get('cliente_doc', '')
        notas = request.form.get('notas', '')
        numero = f'FAC-{v["id"]:06d}'
        fecha_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        try:
            db.execute(
                '''
                INSERT INTO facturas (venta_id, numero, fecha, cliente_nombre, cliente_doc, notas, total)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''',
                (
                    vid,
                    numero,
                    fecha_actual,
                    cliente_nombre,
                    cliente_doc,
                    notas,
                    v['total'],
                ),
            )
            db.commit()
            flash('Factura generada con éxito.', 'success')
        except sqlite3.IntegrityError:
            flash('Esta venta ya fue facturada previamente.', 'error')

        return redirect(url_for('factura_detalle_por_venta', vid=vid))

    return render_template('factura_nueva.html', venta=v)


@app.route('/facturas/detalle-venta/<int:vid>')
@login_required
def factura_detalle_por_venta(vid):
    db = get_db()
    f = db.execute(
        'SELECT * FROM facturas WHERE venta_id = ?', (vid,)
    ).fetchone()
    if not f:
        return redirect(url_for('factura_nueva', vid=vid))
    return redirect(url_for('factura_detalle', fid=f['id']))


@app.route('/facturas/detalle/<int:fid>')
@login_required
def factura_detalle(fid):
    db = get_db()
    f = db.execute('SELECT * FROM facturas WHERE id = ?', (fid,)).fetchone()
    if not f:
        flash('Factura no encontrada.', 'error')
        return redirect(url_for('facturas'))

    items = db.execute(
        'SELECT * FROM venta_items WHERE venta_id = ?', (f['venta_id'],)
    ).fetchall()
    return render_template('factura_detalle.html', factura=f, items=items)


@app.route('/facturas/pdf/<int:fid>')
@login_required
def factura_pdf(fid):
    db = get_db()
    f = db.execute('SELECT * FROM facturas WHERE id = ?', (fid,)).fetchone()
    if not f:
        flash('Factura no encontrada.', 'error')
        return redirect(url_for('facturas'))

    items = db.execute(
        'SELECT * FROM venta_items WHERE venta_id = ?', (f['venta_id'],)
    ).fetchall()
    config = db.execute('SELECT * FROM configuracion WHERE id = 1').fetchone()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    story = []
    styles = getSampleStyleSheet()

    story.append(
        Paragraph(
            f'<b>{config["nombre_negocio"]}</b>', styles['Title']
        )
    )
    if config['cuit']:
        story.append(
            Paragraph(f'CUIT/ID: {config["cuit"]}', styles['Normal'])
        )
    if config['direccion']:
        story.append(
            Paragraph(f'Dirección: {config["direccion"]}', styles['Normal'])
        )

    story.append(Spacer(1, 15))
    story.append(
        Paragraph(
            f'<b>COMPROBANTE N.º: {f["numero"]}</b>', styles['Heading2']
        )
    )
    story.append(Paragraph(f'Fecha: {f["fecha"]}', styles['Normal']))
    story.append(
        Paragraph(f'Cliente: {f["cliente_nombre"]}', styles['Normal'])
    )
    if f['cliente_doc']:
        story.append(
            Paragraph(f'Doc/CUIT: {f["cliente_doc"]}', styles['Normal'])
        )

    story.append(Spacer(1, 15))

    data = [['Producto', 'Cant.', 'P. Unitario', 'Subtotal']]
    for item in items:
        data.append([
            item['nombre_producto'],
            str(item['cantidad']),
            f'${item["precio_unitario"]:,.2f}',
            f'${item["subtotal"]:,.2f}',
        ])

    data.append(['', '', 'TOTAL:', f'${f["total"]:,.2f}'])

    t = Table(data, colWidths=[250, 60, 100, 100])
    t.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.lightgrey),
            ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (2, -1), (-1, -1), 1, colors.black),
        ])
    )

    story.append(t)

    if f['notas']:
        story.append(Spacer(1, 15))
        story.append(
            Paragraph(f'<b>Notas:</b> {f["notas"]}', styles['Normal'])
        )

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'factura_{f["numero"]}.pdf',
        mimetype='application/pdf',
    )


# ==========================================
# USUARIOS Y CONFIGURACIÓN
# ==========================================


@app.route('/usuarios', methods=['GET', 'POST'])
@admin_required
def usuarios():
    db = get_db()
    if request.method == 'POST':
        nombre = request.form['nombre']
        usuario = request.form['usuario']
        password = request.form['password']
        rol = request.form.get('rol', 'vendedor')

        try:
            db.execute(
                'INSERT INTO usuarios (nombre, usuario, password_hash, rol) VALUES (?, ?, ?, ?)',
                (nombre, usuario, generate_password_hash(password), rol),
            )
            db.commit()
            flash('Usuario creado con éxito.', 'success')
        except sqlite3.IntegrityError:
            flash('El nombre de usuario ya existe.', 'error')

        return redirect(url_for('usuarios'))

    u_lista = db.execute(
        'SELECT id, nombre, usuario, rol, activo FROM usuarios ORDER BY id'
    ).fetchall()
    return render_template('usuarios.html', usuarios=u_lista)


@app.route('/usuarios/toggle/<int:uid>', methods=['POST'])
@admin_required
def usuario_toggle(uid):
    if uid == g.user['id']:
        flash('No puedes desactivar tu propio usuario.', 'error')
        return redirect(url_for('usuarios'))

    db = get_db()
    db.execute('UPDATE usuarios SET activo = 1 - activo WHERE id = ?', (uid,))
    db.commit()
    flash('Estado de usuario actualizado.', 'success')
    return redirect(url_for('usuarios'))


@app.route('/configuracion', methods=['GET', 'POST'])
@admin_required
def configuracion():
    db = get_db()
    if request.method == 'POST':
        nombre_negocio = request.form['nombre_negocio']
        cuit = request.form.get('cuit', '')
        direccion = request.form.get('direccion', '')

        db.execute(
            '''
            UPDATE configuracion 
            SET nombre_negocio = ?, cuit = ?, direccion = ? 
            WHERE id = 1
        ''',
            (nombre_negocio, cuit, direccion),
        )
        db.commit()
        flash('Configuración guardada correctamente.', 'success')
        return redirect(url_for('configuracion'))

    config = db.execute('SELECT * FROM configuracion WHERE id = 1').fetchone()
    return render_template('configuracion.html', config=config)


# ==========================================
# EXPORTACIÓN DE EXCEL
# ==========================================


@app.route('/exportar/productos')
@admin_required
def exportar_productos():
    db = get_db()
    prods = db.execute('SELECT * FROM productos ORDER BY nombre').fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    ws.append(
        ['ID', 'Nombre', 'Categoría', 'Costo', 'Precio Venta', 'Stock', 'Stock Mínimo']
    )

    for p in prods:
        ws.append([
            p['id'],
            p['nombre'],
            p['categoria'],
            p['costo'],
            p['precio_venta'],
            p['stock'],
            p['stock_min'],
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='inventario_productos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/exportar/ventas')
@admin_required
def exportar_ventas():
    db = get_db()
    ventas = db.execute('''
        SELECT v.id, v.fecha, u.nombre as vendedor, v.total 
        FROM ventas v 
        LEFT JOIN usuarios u ON v.vendedor_id = u.id 
        ORDER BY v.id DESC
    ''').fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    ws.append(['ID Venta', 'Fecha', 'Vendedor', 'Total'])

    for v in ventas:
        ws.append([v['id'], v['fecha'], v['vendedor'], v['total']])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='reporte_ventas.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/exportar/reporte')
@admin_required
def exportar_reporte():
    db = get_db()
    rango = request.args.get('rango', 'hoy')

    hoy = datetime.now()
    filtro_ventas = ''
    filtro_gastos = ''
    params_v = []
    params_g = []

    if rango == 'hoy':
        inicio = hoy.strftime('%Y-%m-%d 00:00:00')
        fin = hoy.strftime('%Y-%m-%d 23:59:59')
        filtro_ventas = 'WHERE fecha BETWEEN ? AND ?'
        filtro_gastos = 'WHERE fecha BETWEEN ? AND ?'
        params_v = [inicio, fin]
        params_g = [hoy.strftime('%Y-%m-%d'), hoy.strftime('%Y-%m-%d')]
    elif rango == '7':
        filtro_ventas = 'WHERE fecha >= ?'
        filtro_gastos = 'WHERE fecha >= ?'
        params_v = [(hoy - timedelta(days=7)).strftime('%Y-%m-%d 00:00:00')]
        params_g = [(hoy - timedelta(days=7)).strftime('%Y-%m-%d')]
    elif rango == '30':
        filtro_ventas = 'WHERE fecha >= ?'
        filtro_gastos = 'WHERE fecha >= ?'
        params_v = [(hoy - timedelta(days=30)).strftime('%Y-%m-%d 00:00:00')]
        params_g = [(hoy - timedelta(days=30)).strftime('%Y-%m-%d')]

    tot_v = db.execute(
        f'SELECT SUM(total) FROM ventas {filtro_ventas}', params_v
    ).fetchone()[0] or 0

    if filtro_ventas:
        tot_c = db.execute(
            f'SELECT SUM(vi.costo_unitario * vi.cantidad) FROM venta_items vi JOIN ventas v ON vi.venta_id = v.id {filtro_ventas}',
            params_v,
        ).fetchone()[0] or 0
    else:
        tot_c = db.execute(
            'SELECT SUM(costo_unitario * cantidad) FROM venta_items'
        ).fetchone()[0] or 0

    tot_g = db.execute(
        f'SELECT SUM(monto) FROM gastos {filtro_gastos}', params_g
    ).fetchone()[0] or 0
    neto = tot_v - tot_c - tot_g

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen Financiero'

    ws.append(['Métrica', 'Monto'])
    ws.append(['Total Ventas', tot_v])
    ws.append(['Costo Mercadería', tot_c])
    ws.append(['Gastos Operativos', tot_g])
    ws.append(['Ganancia Neta', neto])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'resumen_financiero_{rango}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ==========================================
# INICIO DE LA APLICACIÓN
# ==========================================

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='127.0.0.1', port=5000)
